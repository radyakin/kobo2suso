# Sergiy Radyakin, The World Bank, 2021

import json, os, re, shutil, tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import susoqx

global i
global filemap
global catdict
global _verbose
global varnameslist

i=1
filemap={}
filemap["type"]=-1
filemap["name"]=-1
filemap["text"]=-1
filemap["hint"]=-1
filemap["appearance"]=-1
filemap["calculation"]=-1
catdict={}
_verbose=False
varnameslist=[]


def buildfilemap(ws, lng):
    # scan the first line of the worksheet to determine indices for specific columns
    # lng may be specified or left blank
    # if lng is specified, it must match exactly the suffix after ::
    # if lng is not specified, it will take the first suitable column matching the pattern.
    # this may cause misalignment, such as in case where the columns are interleaved:
    # label::ru, label::fr, hint::fr, hint::ru
    if (lng==None):
        lng=""
    sfx=""
    if (lng!=""):
        sfx="::"+lng

    for j in range(1,255):
      c=ws.cell(column=j, row=1).value
      if (c==None):
          c=""
      if (filemap["type"]==-1 and c=="type"):
          filemap["type"]=j
          print("Located TYPE in column "+str(j))
      if (filemap["name"]==-1 and c=="name"):
          filemap["name"]=j
          print("Located NAME in column "+str(j))
      if (filemap["text"]==-1 and c=="label"+sfx or (c.startswith("label:") and sfx=="")):
          filemap["text"]=j
          print("Located TEXT in column "+str(j))
      if (filemap["hint"]==-1 and c=="hint"+sfx or(c.startswith("hint:") and sfx=="")):
          filemap["hint"]=j
          print("Located HINT in column "+str(j))
      if (filemap["appearance"]==-1 and c=="appearance"):
          filemap["appearance"]=j
          print("Located APPEARANCE in column "+str(j))
      if (filemap["calculation"]==-1 and c=="calculation"):
          filemap["calculation"]=j
          print("Located CALCULATION in column "+str(j))

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False
def is_integer(s):
    if (not is_number(s)):
        return False
    f=float(s)
    if (f!=round(f)):
        return False
    return True

def postcategories(kobofile, choiceset, stagefolder):

    if (_verbose):
      print(" -- Categories: "+choiceset)

    if (choiceset in catdict):
        return catdict[choiceset]
    # new categories, need to export
    f=susoqx.getguid()
    fcopy=f
    f=f.replace("-","")
    fn=os.path.join(stagefolder, "Categories", f+".xlsx")
    writecategories(kobofile, choiceset, fn)
    catdict[choiceset]=fcopy
    return fcopy

def writecategories(kobofile, choiceset, susofile):
  # write kobo choiceset as Survey Solutions categories
  # KOBO: list_name, name, label
  kobo = load_workbook(filename = kobofile, data_only = True)
  kobosheet = kobo["choices"] # must be this specific name according to the format
  # a file has been spotted in the wild where "list_name" was specified as "list name"
  assert(kobosheet.cell(column=1, row=1).value=="list_name" or kobosheet.cell(column=1, row=1).value=="list name")
  assert(kobosheet.cell(column=2, row=1).value=="name")
  assert(
    kobosheet.cell(column=3, row=1).value=="label"
    or
    kobosheet.cell(column=3, row=1).value[0:7]=="label::"
  )
  if (_verbose):
    print(" -- Writing to: "+susofile)
  suso=Workbook()
  susosheet = suso.active
  susosheet.title = "Categories" # must be this specific name according to the format
  susosheet.cell(column=1,row=1).value="id"
  susosheet.cell(column=2,row=1).value="text"
  susosheet.cell(column=3,row=1).value="parentid"
  susosheet.column_dimensions[get_column_letter(2)].width=60
  susorow=2
  koborow=2
  empty=0
  # todo: currently no checks are done to ensure no overlap
  #       of these new codes with codes in the choiceset
  newcode=100001
  while(empty<10):  # more than 2 empty lines occur commonly in the files spotted in the wild
      listname=kobosheet.cell(column=1,row=koborow).value
      name    =kobosheet.cell(column=2,row=koborow).value
      label   =kobosheet.cell(column=3,row=koborow).value
      if (listname==None):
          listname=""
      if(listname==""):
          empty=empty+1
      else:
          empty=0
          # print(listname)
          if (str(listname).strip()==str(choiceset).strip()):
              #print(name, label)
              if (not is_integer(name)):
                  label=label+" ["+name+"]"
                  name=newcode
                  newcode=newcode+1
              susosheet.cell(column=1,row=susorow).value=name
              susosheet.cell(column=2,row=susorow).value=label   ## ToDo: need to trim this to 250 chars
              susorow=susorow+1
      koborow=koborow+1
  suso.save(filename = susofile)


def getkoboline(ws,i):
    kobo={}

    for t in ["type", "name", "text", "hint", "appearance", "calculation"]:
        kobo[t]=""
        if (filemap[t]>0):
            kobo[t]=ws.cell(column=filemap[t],row=i).value
            if (kobo[t]==None):
                kobo[t]=""

    # todo: this is a good place to replace alternative spellings
    kobo['type']=re.sub("^begin group", "begin_group", kobo['type'])
    kobo['type']=re.sub("^end group", "end_group", kobo['type'])

    kobo['type']=re.sub("^select one from ", "select_one ", kobo['type'])
    kobo['type']=re.sub("^select one ", "select_one ", kobo['type'])
    kobo['type']=re.sub("^select1 ", "select_one ", kobo['type'])
    kobo['type']=re.sub("^select all that apply from ", "select_multiple ", kobo['type'])
    kobo['type']=re.sub("^select all that apply ", "select_multiple ", kobo['type'])

    ########## THE FOLLOWING CONVERTS ROSTERS INTO SUB-SECTIONS ################
    kobo['type']=re.sub("^begin repeat", "begin_group", kobo['type'])
    kobo['type']=re.sub("^end repeat", "end_group", kobo['type'])
    ############################################################################

    kobosplit=kobo['type'].split()
    kobo['type1']=""
    if (len(kobosplit)>0):
      kobo['type1']=kobosplit[0]
    kobo['type2']=""
    if (len(kobosplit)>1):
      kobo['type2']=kobosplit[1]

    return kobo


def processgroup(kobofile, ws, name, title, stagefolder):
  global i
  global varnameslist

  G=susoqx.getgroup(title) # // todo: must also pass name
  C=[]

  empty=0
  while (empty<2):
    kobo=getkoboline(ws,i)
    i=i+1

    kobo['text']=susoqx.adaptsubst(kobo['text'])
    kobo['hint']=susoqx.adaptsubst(kobo['hint'])

    if (kobo['type']==None or kobo['type']==""):
      if (empty==1):
        break
      else:
        empty=empty+1
    else:
      empty=0
      if (kobo['name']!="" and kobo['name']!=None):
          varnameslist.append(kobo['name'])

      pfx="   "
      if (kobo['type1']=="begin_group"):
        print("------------------------------------------------------------")
        pfx=""

      tag=kobo['type1']
      print(pfx + tag.ljust(20) + "  " + kobo['name'].ljust(28) + "  " + kobo['text'])

      if (tag=="end_group"):
        # a file has been spotted in the wild where "end_group" is not accompanied by the group name
        break # // end of current group

      if (tag=="begin_group"):
        # // encountered a nested group
        # print(kobo['name'])
        Z=processgroup(kobofile, ws, kobo['name'], kobo['text'], stagefolder)
        C.append(Z)

      if (tag=="note"):
        # // note maps to static text
        # Display a note on the screen, takes no input.
        T=susoqx.gettext(kobo['text'])
        C.append(T)

      if (tag in ["start","end","today","deviceid","imei","username","simserial","subscriberid","phonenumber","audit"]):
          msg=susoqx.getmetatext(tag)
          T=susoqx.gettext(msg)
          T['ConditionExpression']="false" # must be string - this is C# syntax
          T['HideIfDisabled']=True
          C.append(T)

      if (tag=="select_one" or tag=="select_multiple"):
        if (kobo['type2']==""):
            print("Error! Expected categories name for "+kobo['name'])
            return
        selectQ=susoqx.getquestion(kobo)
        selectQ['CategoriesId']=postcategories(kobofile,kobo['type2'],stagefolder)
        C.append(selectQ)

      if (tag=="calculate"):
          KV=susoqx.getvar(kobo)
          C.append(KV)

      if (tag in ["text", "integer", "decimal", "date", "barcode", "image", "audio", "geopoint"]):
        C.append(susoqx.getquestion(kobo))

      if (not(tag in ["end_group", "begin_group", "note", "text", "calculate",
      "integer", "decimal", "select_one", "select_multiple", "date", "barcode", "image", "audio", "geopoint",
      "audit", "phonenumber", "subscriberid", "simserial", "username", "deviceid", "imei", "today", "end", "start"
      ])):
        print("!  >>>>>> Encountered an unknown type: "+tag+", skipping")

  G['Children']=C
  return G

def koboConvert(koboname, susoname):
  print("Converter of questionnaires from XFORMS [Kobo Toolbox, ODK, SurveyCTO, Survey123] to Survey Solutions")
  print("2021 Sergiy Radyakin, The World Bank")
  print("Version 0.1")
  print("-----------")

  # todo: for xls files consider https://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx
  global i
  print(koboname)
  wb = load_workbook(filename = koboname, data_only = True)
  ws = wb["survey"]
  print(ws.title)

  buildfilemap(ws, "") # for the moment always take first language

  i=1
  if (ws.cell(column=filemap["type"],row=i).value!="type"):
    print(">>> ERROR")
    return
  print(">>> OK")
  i=i+1

  with tempfile.TemporaryDirectory() as tmpdirname:
      print('created temporary directory', tmpdirname)
      os.mkdir(os.path.join(tmpdirname,"Categories"))
      qxdoc=susoqx.getqx("Q")
      G=processgroup(koboname, ws, "Main", "MAIN", tmpdirname)
      print("Finished reading. Stopped at line ",i," of ",koboname)
      qxdoc['Children'].append(G)
      C=[]
      for key, value in catdict.items():
          pair={}
          pair["Id"]  =value
          pair["Name"]="suso_"+key
          C.append(pair)
      qxdoc['Categories']=C

      with open(os.path.join(tmpdirname, "document.json"), 'w') as outfile:
        json.dump(qxdoc, outfile, indent=2)

      shutil.make_archive(susoname,"zip",tmpdirname)

      unique=[]
      dups=[]
      for varname in varnameslist:
        if varname not in unique:
          unique.append(varname)
        else:
          dups.append(varname)

      if (len(dups)>0):
        print("Duplicates")
        for varname in dups:
          print(varname)
  # END OF FILE
